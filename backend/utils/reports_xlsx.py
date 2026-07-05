"""
Phase 7 — Excel report builder.

Every workbook ships with a "Data" sheet holding the actual cleaned
transaction rows as a named Excel Table, plus a "Summary" sheet with the
same figures shown in the PDF version of the same report. Where a figure
can be verified with a genuine live formula against the Data table
(total revenue, average order value) we add a "(verify)" formula cell
next to it — real spreadsheet dynamism — but the authoritative reported
numbers are written as values so the Excel and PDF versions of a report
never disagree with each other.
"""
import io
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

NAVY = "101728"
TEAL = "0F9E7A"
LIGHT = "F5F7FB"
WHITE = "FFFFFF"

HEADER_FONT = Font(color=WHITE, bold=True)
HEADER_FILL = PatternFill("solid", start_color=NAVY, end_color=NAVY)
TITLE_FONT = Font(color=NAVY, bold=True, size=14)
LABEL_FONT = Font(color="5B6478")
NOTE_FONT = Font(color="5B6478", italic=True, size=9)


def _safe_table_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", name)
    if not cleaned or not cleaned[0].isalpha():
        cleaned = f"T_{cleaned}"
    return cleaned[:60]


def _write_data_sheet(wb, df, table_name):
    ws = wb.create_sheet("Data")
    headers = [str(c) for c in df.columns]
    ws.append(headers)
    for _, row in df.iterrows():
        ws.append([row[c] if row[c] == row[c] else None for c in df.columns])  # NaN-safe

    n_rows, n_cols = df.shape
    last_col_letter = get_column_letter(n_cols)
    ref = f"A1:{last_col_letter}{n_rows + 1}"

    table = Table(displayName=_safe_table_name(table_name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False
    )
    ws.add_table(table)

    for col_idx, header in enumerate(headers, start=1):
        width = max(12, min(28, len(str(header)) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    return ws, table.displayName


def _new_summary_sheet(wb, title):
    ws = wb.create_sheet("Summary")
    ws["A1"] = "INSIGHT IQ"
    ws["A1"].font = Font(color=TEAL, bold=True, size=10)
    ws["A2"] = title
    ws["A2"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    return ws


def _write_meta(ws, row, dataset_filename, row_count, source_stage, generated_at):
    ws.cell(row=row, column=1, value=f"Dataset: {dataset_filename}").font = LABEL_FONT
    ws.cell(row=row + 1, column=1, value=f"{row_count:,} rows ({source_stage} data)").font = LABEL_FONT
    ws.cell(row=row + 2, column=1, value=f"Generated {generated_at}").font = LABEL_FONT
    return row + 4


def _section_title(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(color=NAVY, bold=True, size=12)
    return row + 1


def _write_table_block(ws, row, headers, rows, col_widths=None):
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1
    start_data_row = row
    for r_idx, data_row in enumerate(rows):
        for c, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=c, value=value)
            if (row - start_data_row) % 2 == 1:
                cell.fill = PatternFill("solid", start_color=LIGHT, end_color=LIGHT)
        row += 1
    if col_widths:
        for c, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = width
    return row + 1


def _kv_rows(ws, row, pairs):
    for label, value in pairs:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=value)
        row += 1
    return row + 1


def _verify_formula_row(ws, row, label, formula):
    ws.cell(row=row, column=1, value=label).font = NOTE_FONT
    cell = ws.cell(row=row, column=2, value=formula)
    cell.font = NOTE_FONT
    return row + 1


# ---------------------------------------------------------------------------
# Per-report-type sheet builders
# ---------------------------------------------------------------------------


def _executive_summary_sheet(ws, row, context, table_name, field_matches):
    row = _section_title(ws, row, "Key Business Metrics")
    kpis = context["kpis"]
    row = _kv_rows(
        ws,
        row,
        [
            ("Total Revenue", kpis.get("total_revenue")),
            ("Total Orders", kpis.get("total_orders")),
            ("Total Customers", kpis.get("total_customers")),
            ("Profit Margin (%)", kpis.get("profit_margin_percent")),
            ("Average Order Value", kpis.get("average_order_value")),
            ("Monthly Growth (%)", kpis.get("monthly_growth_percent")),
        ],
    )
    revenue_col = field_matches.get("revenue")
    if revenue_col:
        row = _verify_formula_row(ws, row, "Total Revenue (verify, live formula)", f"=SUM({table_name}[{revenue_col}])")

    row = _section_title(ws, row + 1, "Highlights")
    if context["has_insights"] and context["highlights"]:
        for h in context["highlights"]:
            ws.cell(row=row, column=1, value=f"• {h}")
            row += 1
    else:
        ws.cell(row=row, column=1, value="Run Business Insights on this dataset to populate highlights here.").font = NOTE_FONT
        row += 1

    row = _section_title(ws, row + 1, "Top Performers")
    if context["top_product"]:
        keys = list(context["top_product"].keys())
        ws.cell(row=row, column=1, value=f"Top product: {context['top_product'][keys[0]]}")
        row += 1
    if context["top_region"]:
        ws.cell(row=row, column=1, value=f"Top region: {context['top_region']['name']} ({context['top_region']['value']:,.0f})")
        row += 1
    return row


def _sales_sheet(ws, row, context):
    trend = context["revenue_trend"]
    row = _section_title(ws, row, "Revenue Trend")
    if trend:
        rows = list(zip(trend["x"], trend["y"]))
        row = _write_table_block(ws, row, ["Month", "Revenue"], rows)
    else:
        ws.cell(row=row, column=1, value="No order date/revenue column detected.").font = NOTE_FONT
        row += 2

    regional = context["regional_sales"]
    row = _section_title(ws, row, "Regional Sales")
    if regional:
        rows = list(zip(regional["labels"], regional["values"]))
        row = _write_table_block(ws, row, ["Region", "Revenue"], rows)
    else:
        ws.cell(row=row, column=1, value="No region column detected.").font = NOTE_FONT
        row += 2

    for title, key in [("Top Selling Products", "top_products"), ("Low Performing Products", "low_products")]:
        row = _section_title(ws, row, title)
        items = context[key]
        if items:
            keys = list(items[0].keys())
            headers = [k.replace("_", " ").title() for k in keys]
            rows = [[p[k] for k in keys] for p in items]
            row = _write_table_block(ws, row, headers, rows)
        else:
            ws.cell(row=row, column=1, value="No product column detected.").font = NOTE_FONT
            row += 2

    return row


def _customer_analytics_sheet(ws, row, context):
    row = _section_title(ws, row, "Customer Overview")
    row = _kv_rows(
        ws,
        row,
        [
            ("Total Customers", context["total_customers"]),
            ("Average Order Value", context["average_order_value"]),
        ],
    )

    row = _section_title(ws, row, "Top Customers by Revenue")
    dist = context["customer_distribution"]
    if dist:
        rows = list(zip(dist["labels"], dist["values"]))
        row = _write_table_block(ws, row, ["Customer", "Total Revenue"], rows)
    else:
        ws.cell(row=row, column=1, value="No customer column detected.").font = NOTE_FONT
        row += 2

    row = _section_title(ws, row, "Customer Segments")
    if context["segmentation"]:
        rows = [
            [s["label"], s["size"], s["avg_total_revenue"], s["avg_order_count"]]
            for s in context["segmentation"]["segments"]
        ]
        row = _write_table_block(ws, row, ["Segment", "Customers", "Avg Revenue", "Avg Orders"], rows)
    else:
        note = "Run Predictive Analytics to include customer segments." if not context["predictive_available"] else "Not enough customers to segment."
        ws.cell(row=row, column=1, value=note).font = NOTE_FONT
        row += 2

    row = _section_title(ws, row, "Churn Risk")
    if context["churn"]:
        ws.cell(row=row, column=1, value=f"Estimated churn rate: {context['churn']['churn_rate_percent']}%")
        row += 2
        rows = [[c["customer"], c["churn_probability"], c["recency_days"]] for c in context["churn"]["at_risk_customers"][:10]]
        row = _write_table_block(ws, row, ["Customer", "Churn Probability", "Days Since Last Order"], rows)
    else:
        note = "Run Predictive Analytics to include churn risk." if not context["predictive_available"] else "Not enough history to estimate churn."
        ws.cell(row=row, column=1, value=note).font = NOTE_FONT
        row += 2

    return row


def _financial_sheet(ws, row, context, table_name, field_matches):
    row = _section_title(ws, row, "Financial Summary")
    row = _kv_rows(
        ws,
        row,
        [
            ("Total Revenue", context["total_revenue"]),
            ("Profit Margin (%)", context["profit_margin_percent"]),
            ("Average Order Value", context["average_order_value"]),
            ("Monthly Growth (%)", context["monthly_growth_percent"]),
        ],
    )
    revenue_col = field_matches.get("revenue")
    if revenue_col:
        row = _verify_formula_row(ws, row, "Total Revenue (verify, live formula)", f"=SUM({table_name}[{revenue_col}])")

    row = _section_title(ws, row + 1, "Revenue vs. Profit")
    rvp = context["revenue_vs_profit"]
    if rvp:
        rows = list(zip(rvp["x"], rvp["revenue"], rvp["profit"]))
        row = _write_table_block(ws, row, ["Month", "Revenue", "Profit"], rows)
    else:
        ws.cell(row=row, column=1, value="No profit column and/or order date detected.").font = NOTE_FONT
        row += 2

    return row


def _inventory_sheet(ws, row, context):
    inv = context["inventory_status"]
    row = _section_title(ws, row, "Inventory Overview")
    if inv:
        row = _kv_rows(
            ws,
            row,
            [
                ("Total Units In Stock", inv["total_units_in_stock"]),
                ("Low-Stock Products", inv["low_stock_count"]),
            ],
        )
        row = _section_title(ws, row, "Low Stock Items")
        if inv["low_stock_items"]:
            rows = [[i["name"], i["stock"]] for i in inv["low_stock_items"]]
            row = _write_table_block(ws, row, ["Product", "Stock Remaining"], rows)
        else:
            ws.cell(row=row, column=1, value="No products are below the low-stock threshold.")
            row += 2
    else:
        ws.cell(row=row, column=1, value="No stock/inventory column detected.").font = NOTE_FONT
        row += 2

    if context.get("inventory_insight"):
        row = _section_title(ws, row, "Note")
        ws.cell(row=row, column=1, value=context["inventory_insight"])
        row += 2

    return row


_SHEET_BUILDERS = {
    "executive-summary": _executive_summary_sheet,
    "sales": _sales_sheet,
    "customer-analytics": _customer_analytics_sheet,
    "financial": _financial_sheet,
    "inventory": _inventory_sheet,
}
_NEEDS_TABLE_CONTEXT = {"executive-summary", "financial"}


def build_xlsx(context: dict, df, field_matches: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    data_ws, table_name = _write_data_sheet(wb, df, f"{context['report_type']}_data")

    summary_ws = _new_summary_sheet(wb, context["report_title"])
    row = _write_meta(summary_ws, 3, context["dataset_filename"], context["row_count"], context["source_stage"], context["generated_at"])

    builder = _SHEET_BUILDERS[context["report_type"]]
    if context["report_type"] in _NEEDS_TABLE_CONTEXT:
        builder(summary_ws, row, context, table_name, field_matches)
    else:
        builder(summary_ws, row, context)

    wb.move_sheet("Summary", offset=-len(wb.sheetnames))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
